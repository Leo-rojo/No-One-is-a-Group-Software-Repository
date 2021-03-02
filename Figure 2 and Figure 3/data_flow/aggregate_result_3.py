# importing openpyxl module
import openpyxl as xl
import os
import glob
import csv
from shutil import copyfile

datain='C:\\Users\\leona\\Desktop\\No-One-is-a-Group-Software-Repository\\Figure 2 and Figure 3\\data_flow\\data\\resultsmixture\\'

#convert csv files to xlsx files
os.chdir(datain)
print(os.getcwd())
FileList = glob.glob('*.csv')

for file in FileList:
    wb = xl.Workbook()
    ws = wb.active

    with open(file) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    wb.save(file+'.xlsx')

# merge userresults into same file
FileList = glob.glob('*.xlsx')
offj=0
for file in FileList:
    wb1 = xl.load_workbook(datain+file)
    ws1 = wb1.worksheets[0]

    # opening the destination excel file
    filename1 = "C:\\Users\\leona\\Desktop\\No-One-is-a-Group-Software-Repository\\Figure 2 and Figure 3\\data_flow\\data\\final_results.xlsx"
    wb2 = xl.load_workbook(filename1)
    ws2 = wb2.active

    # calculate total number of rows and
    # columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column

    # copying the cell values from source
    # excel file to destination excel file
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row=i, column=j)

            # writing the read value to destination excel file
            ws2.cell(row=i, column=j+offj).value = c.value
    ws2.cell(row=mr+1,column=offj+1).value=file

        # saving the destination excel file
    offj=offj+5
    wb2.save(str(filename1))

os.chdir('..')
os.chdir('..')
#add mos results
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active
for v in range(5,160,5):
    for i in range(16, 28):
        for j in range(1, 5):
            ws2.cell(row=i,column=v+j).value = ws2.cell(row=i,column=j).value
wb2.save('.\\data\\final_results_with_mos.xlsx')

#calculate difference respect to mos
wb2 = xl.load_workbook('.\\data\\final_results_with_mos.xlsx')
ws2 = wb2.active
for i in range(1,13):
    for j in range(1,160):
        try:
            ws2.cell(row=i+30,column=j).value = abs(float(ws2.cell(row=i+15,column=j).value))-abs(float(ws2.cell(row=i,column=j).value))
        except:
            print('empty cell or not convertible cell')
wb2.save('.\\data\\final_results_with_mos_and_difference.xlsx')