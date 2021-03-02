import openpyxl
import csv
from openpyxl.utils import get_column_letter
from six import string_types
import numpy as np
import matplotlib.pyplot as plt

modelli=['Exponential','Linear Rebuffering','Linear Utility','IIR filter','MPC','Bola','SDNDASH','SQI','P1203','VideoATLAS','KSQI']
wb1 = openpyxl.load_workbook('.\\data\\final_results_with_mos_and_difference.xlsx')
ws=wb1.active

##no diff ma valori reali

plcc = []
srcc = []
krcc = []
contain11plcc=[]
contain11srcc=[]
contain11krcc=[]

# real value
for riga in range(2,13):  # ,43):
    # pcc
    for col in range(2, 160, 5):
        plcc.append(float(ws.cell(row=riga, column=col).value))
    # srcc
    for col in range(3, 160, 5):
        srcc.append(float(ws.cell(row=riga, column=col).value))
    # krcc
    for col in range(4, 160, 5):
        krcc.append(float(ws.cell(row=riga, column=col).value))
    contain11plcc.append(plcc)
    contain11srcc.append(srcc)
    contain11krcc.append(krcc)
    plcc = []
    srcc = []
    krcc = []

# plot real value
fig1 = plt.figure(figsize=(6,3.2))
count=0
for i in contain11plcc:
    model=modelli[count]
    surt = np.sort(i)
    p = 1. * np.arange(len(i)) / (len(i) - 1)
    # plt.step(sortedtime, p)
    plt.step(surt, p,label=model)
    plt.axvline(x=0.5,linestyle='--',color='k')
    plt.axvline(x=-0.5,linestyle='--',color='k')
    leg=plt.legend(loc='upper left', bbox_to_anchor=(0.05, 1.20),
      ncol=6,prop={'size': 4})#'weight':'bold'
    leg.get_frame().set_linewidth(0.0)
    leg.get_frame().set_edgecolor('k')
    plt.xlabel('PCC',fontsize=15,fontweight='bold')
    plt.ylabel('Fraction of users',fontsize=15,fontweight='bold')
    count+=1
fig1.savefig(".\\data\\plots\\plcc_hdtv.pdf", bbox_inches='tight')

fig2 = plt.figure(figsize=(6,3.2))

count=0
for i in contain11srcc:
    model = modelli[count]
    surt = np.sort(i)
    p = 1. * np.arange(len(i)) / (len(i) - 1)
    # plt.step(sortedtime, p)
    plt.step(surt, p,label=model)
    plt.axvline(x=0.5, linestyle='--', color='k')
    plt.axvline(x=-0.5, linestyle='--', color='k')
    #plt.legend(loc='upper left', bbox_to_anchor=(0.05, 1.20),
    #           ncol=4, prop={'size': 6, 'weight': 'bold'})
    plt.xlabel('SRCC', fontsize=15, fontweight='bold')
    plt.ylabel('Fraction of users', fontsize=15, fontweight='bold')
    count += 1
fig2.savefig(".\\data\\plots\\srcc_hdtv.pdf", bbox_inches='tight')

fig3 = plt.figure(figsize=(6,3.2))
count=0
for i in contain11krcc:
    model = modelli[count]
    surt = np.sort(i)
    p = 1. * np.arange(len(i)) / (len(i) - 1)
    # plt.step(sortedtime, p)
    plt.step(surt, p,label=model)
    plt.axvline(x=0.5, linestyle='--', color='k')
    plt.axvline(x=-0.5, linestyle='--', color='k')
    #plt.legend(loc='upper left', bbox_to_anchor=(0.05, 1.20),
    #           ncol=4, prop={'size': 6, 'weight': 'bold'})
    plt.xlabel('KRCC', fontsize=15, fontweight='bold')
    plt.ylabel('Fraction of users', fontsize=15, fontweight='bold')
    count += 1
fig3.savefig(".\\data\\plots\\krcc_hdtv.pdf", bbox_inches='tight')
plt.show()





##diffvalues
diffplcc=[]
diffsrcc=[]
diffkrcc=[]
diffcontain11plcc=[]
diffcontain11srcc=[]
diffcontain11krcc=[]
#cella riferimento
for riga in range(32,43):#,43):
    #pcc
    for col in range(2,160,5):
        diffplcc.append(ws.cell(row=riga,column=col).value)
    #srcc
    for col in range(3,160,5):
        diffsrcc.append(ws.cell(row=riga,column=col).value)
    #krcc
    for col in range(4,160,5):
        diffkrcc.append(ws.cell(row=riga,column=col).value)

    diffcontain11plcc.append(diffplcc)
    diffcontain11srcc.append(diffsrcc)
    diffcontain11krcc.append(diffkrcc)
    diffplcc = []
    diffsrcc = []
    diffkrcc = []

count=0
fig1 = plt.figure(figsize=(6,3.2))
for i in diffcontain11plcc:
    model = modelli[count]
    surt = np.sort(i)
    p = 1. * np.arange(len(i))/(len(i) - 1)
    #plt.step(sortedtime, p)
    plt.step(surt,p,label=model)
    plt.axvline(x=0,linestyle='--',color='k')
    #plt.legend(loc='upper left', bbox_to_anchor=(0.05, 1.20),
    #           ncol=4, prop={'size': 6, 'weight': 'bold'})
    plt.xlabel('PCC magnitude difference', fontsize=15, fontweight='bold')
    plt.ylabel('Fraction of users', fontsize=15, fontweight='bold')
    count += 1
fig1.savefig(".\\data\\plots\\diffplcc_hdtv.pdf", bbox_inches='tight')

count=0
fig2 = plt.figure(figsize=(6,3.2))
for i in diffcontain11srcc:
    model = modelli[count]
    surt = np.sort(i)
    p = 1. * np.arange(len(i))/(len(i) - 1)
    #plt.step(sortedtime, p)
    plt.step(surt,p,label=model)
    plt.axvline(x=0,linestyle='--',color='k')
    #plt.legend(loc='upper left', bbox_to_anchor=(0.05, 1.20),
    #           ncol=4, prop={'size': 6, 'weight': 'bold'})
    plt.xlabel('SRCC magnitude difference', fontsize=15, fontweight='bold')
    plt.ylabel('Fraction of users', fontsize=15, fontweight='bold')
    count+=1
fig2.savefig(".\\data\\plots\\diffsrcc_hdtv.pdf", bbox_inches='tight')

count=0
fig3=plt.figure(figsize=(6,3.2))
for i in diffcontain11krcc:
    model = modelli[count]
    surt = np.sort(i)
    p = 1. * np.arange(len(i))/(len(i) - 1)
    #plt.step(sortedtime, p)
    plt.step(surt,p,label=model)
    plt.axvline(x=0,linestyle='--',color='k')
    #plt.legend(loc='upper left', bbox_to_anchor=(0.05, 1.20),
    #           ncol=4, prop={'size': 6, 'weight': 'bold'})
    plt.xlabel('KRCC magnitude difference', fontsize=15, fontweight='bold')
    plt.ylabel('Fraction of users', fontsize=15, fontweight='bold')
    count += 1
plt.show()
fig3.savefig(".\\data\\plots\\diffkrcc_hdtv.pdf", bbox_inches='tight')